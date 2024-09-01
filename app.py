import tkinter as tk
from tkinter import filedialog
import os
from flask import Flask, jsonify, render_template, request
from openpyxl import load_workbook
from datetime import datetime
import time
from threading import Thread
import socket
import webbrowser
import openpyxl

# Flask app setup
app = Flask(__name__)

COLOR_CODES = {
    'Red': 'FFFF0000',
    'Blue': 'FF00B0F0',
    'Green': 'FF00B050',
    'Normal': None
}

REVERSED_COLOR_CODES = {
    'FFFF0000': 'Red',
    'FF00B0F0': 'Blue',
    'FF00B050': 'Green',
}

directory = None  # Global variable to store directory path
cached_data = []  # Global variable to store cached data

def get_file_path():
    if directory:
        path = os.path.join(directory, 'DAILY JOB 2023.xlsx')
        print(f"path: {path}")
        return path
    return None

def get_cell_color(cell):
    """Extract color information from a cell's font."""
    if cell.font and cell.font.color and cell.font.color.type == 'rgb':
        return cell.font.color.rgb
    return None

def read_excel_rows(filepath, fromrow, torow):
    """
    Read and filter rows from an Excel file based on the color of the cell in column D.
    """
    if not filepath:
        raise ValueError("File path must be provided")

    data = []

    try:
        # Open the workbook in read-only mode
        workbook = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet = workbook.active
        
        # Process rows
        for row in sheet.iter_rows(min_row=fromrow, max_row=torow, values_only=False):
            cell = row[3]  # Get the cell in column D (index 3)
            color = get_cell_color(cell)
            cell_data = [c.value for c in row]

            if color and cell_data[3]:
                color_name = REVERSED_COLOR_CODES.get(color, "Normal")
                cell_data.append(color_name)
                
                if color in COLOR_CODES.values():
                    data.append(cell_data)
            else:
                cell_data.append("Normal")

        if not data:
            print("Warning: No data found in the specified row range.")
    except Exception as e:
        raise RuntimeError(f"Failed to read from workbook: {e}")
    finally:
        # Ensure resources are cleaned up
        workbook.close()  # This is generally recommended to ensure resources are freed

    return data



def refresh_data(interval=30):
    """Background thread to refresh data every `interval` seconds."""
    global cached_data
    last_mod_time = None
    while True:
        try:
            file_path = get_file_path()
            if file_path:
                current_mod_time = os.path.getmtime(file_path)
                if last_mod_time is None or current_mod_time != last_mod_time:
                    print(f"Refreshing data from {file_path}...")
                    try:
                        cached_data = read_excel_rows(file_path, fromrow=4, torow=100000)
                    except RuntimeError as e:
                        print(f"Error reading file: {e}")
                    last_mod_time = current_mod_time
                else:
                    print("No changes detected in the file.")
            else:
                print("No file path available to refresh data.")
        except Exception as e:
            print(f"Error refreshing data: {e}")
        time.sleep(interval)

def parse_date(date_str):
    if date_str is None:
        return datetime.min  # Use a minimal date for comparison
    formats = ["%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y"]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    raise ValueError(f"Unsupported date format: {date_str}")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/list')
def list_entries():
    global cached_data

    fromdate_str = request.args.get('fromdate')
    todate_str = request.args.get('todate')

    fromdate = parse_date(fromdate_str) if fromdate_str else None
    todate = parse_date(todate_str) if todate_str else None

    combined_list = cached_data

    if fromdate or todate:
        filtered_list = []
        for entry in combined_list:
            entry_date = parse_date(entry[1])
            if entry_date:
                if fromdate and todate:
                    if fromdate <= entry_date <= todate:
                        filtered_list.append(entry)
                elif fromdate:
                    if fromdate <= entry_date:
                        filtered_list.append(entry)
                elif todate:
                    if entry_date <= todate:
                        filtered_list.append(entry)
        combined_list = filtered_list

    combined_list.sort(key=lambda x: parse_date(x[1]))

    refresh_interval = int(request.args.get('interval', 10))  # Default to 30 seconds
    entries_per_set = int(request.args.get('entries', 12))  # Default to 12 entries

    list_length = len(combined_list)

    if list_length > 10:
        current_time = int(time.time())
        offset = (current_time // refresh_interval) % max(list_length, 1)

        rotated_list = combined_list[offset:offset + entries_per_set]

        if len(rotated_list) < entries_per_set:
            rotated_list += combined_list[:entries_per_set - len(rotated_list)]

        return jsonify(rotated_list)
    else:
        return jsonify(combined_list)

@app.route('/list2')
def list2data():
    return render_template("list2.html")

@app.route('/stop_server', methods=['POST'])
def stop_server():
    print("Shutting down the server...")
    os._exit(0)  

def is_port_in_use(port):
    """Check if the port is already in use."""
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex(('localhost', port)) == 0

def start_flask_server():
    if not is_port_in_use(5000):
        app.run(port=5000)
    else:
        print("Flask server is already running on port 5000. Aborting...")

def select_file():
    global directory, cached_data
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    
    if file_path:
        directory = os.path.dirname(file_path)
        result_label.config(text=f"Selected File Directory: {directory}")
        root.destroy()
        cached_data = read_excel_rows(file_path, fromrow=4, torow=100000)
        # Open the Flask app in the default web browser
        webbrowser.open("http://localhost:5000")

root = tk.Tk()
root.title("File Directory Finder")
root.geometry("400x200")

root.configure(bg="#f0f4f8")

frame = tk.Frame(root, bg="#ffffff", padx=20, pady=20, borderwidth=2, relief="groove")
frame.pack(padx=10, pady=10, expand=True, fill=tk.BOTH)

open_button = tk.Button(
    frame,
    text="Select File",
    command=select_file,
    bg="#4CAF50",
    fg="#ffffff",
    font=("Arial", 14),
    padx=10,
    pady=5,
    relief="raised",
    borderwidth=2
)
open_button.pack(pady=10)

result_label = tk.Label(
    frame,
    text="Selected File Directory:",
    bg="#ffffff",
    fg="#333333",
    font=("Arial", 12),
    wraplength=350
)
result_label.pack(pady=10)

# Start the background thread to refresh data
data_refresh_thread = Thread(target=refresh_data, daemon=True)
data_refresh_thread.start()

# Run the Flask server in a separate thread
flask_thread = Thread(target=start_flask_server)
flask_thread.start()

# Run the Tkinter application
root.mainloop()
